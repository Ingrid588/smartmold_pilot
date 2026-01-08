"""
SmartMold Pilot V3 - Excel Data Parser
解析用户上传的Excel测试数据文件

支持的数据格式:
1. 粘度曲线数据 (Step 1)
2. 型腔平衡数据 (Step 2)
3. 压力降数据 (Step 3)
4. 工艺窗口数据 (Step 4)
5. 浇口冻结数据 (Step 5)
6. 冷却时间数据 (Step 6)
7. 锁模力数据 (Step 7)
"""

import os
import io
from typing import Dict, Any, List, Optional, Tuple
from dataclasses import dataclass, field
import openpyxl
from openpyxl import Workbook
from datetime import datetime


@dataclass
class ViscosityData:
    """Step 1: 粘度曲线数据 - 存储原始测量值"""
    speed_percents: List[float] = field(default_factory=list)  # 速度百分比 (%)
    speed_mm_s: List[float] = field(default_factory=list)  # 实际速度 (mm/s)
    switch_positions: List[float] = field(default_factory=list)  # V/P切换位置 (mm)
    fill_times: List[float] = field(default_factory=list)  # 填充时间 (s)
    peak_pressures: List[float] = field(default_factory=list)  # 峰值压力 (Bar)
    screw_diameter: float = 0.0  # mm
    material: str = ""
    machine: str = ""


@dataclass
class CavityBalanceData:
    """Step 2: 型腔平衡数据 - 区分短射和满射"""
    cavity_weights: Dict[int, float] = field(default_factory=dict)  # 腔号 -> 重量(g)
    visual_checks: Dict[int, str] = field(default_factory=dict)     # 腔号 -> OK/NG
    test_type: str = ""  # "Short_Shot" 或 "VP_Switch"
    injection_speed: float = 0.0  # mm/s (从Step1继承)


@dataclass
class PressureDropData:
    """Step 3: 压力降数据"""
    positions: List[str] = field(default_factory=list)  # 位置名称
    pressures: List[float] = field(default_factory=list)  # MPa
    injection_speed: float = 0.0


@dataclass
class ProcessWindowData:
    """Step 4: 工艺窗口数据"""
    speeds: List[float] = field(default_factory=list)  # mm/s
    pressures: List[float] = field(default_factory=list)  # MPa
    product_weights: List[float] = field(default_factory=list)  # Product Weight (g)
    hold_times: List[float] = field(default_factory=list)  # 保压时间 (s)
    quality_ok: List[bool] = field(default_factory=list)  # 是否合格


@dataclass
class GateFreezeData:
    """Step 5: 浇口冻结数据"""
    hold_times: List[float] = field(default_factory=list)  # 秒
    weights: List[float] = field(default_factory=list)  # 克


@dataclass
class CoolingTimeData:
    """Step 6: 冷却时间数据"""
    cooling_times: List[float] = field(default_factory=list)  # 秒
    part_temps: List[float] = field(default_factory=list)  # °C
    deformations: List[float] = field(default_factory=list)  # mm


@dataclass
class ClampingForceData:
    """Step 7: 锁模力数据"""
    forces: List[float] = field(default_factory=list)  # 吨
    part_weights: List[float] = field(default_factory=list)  # 产品重量 (g)
    flash_detected: List[bool] = field(default_factory=list)  # 是否有飞边


@dataclass
class MachineSnapshotData:
    """项目综合信息 - 扩充版（对应MIL标准）"""
    # 产品信息
    model_no: str = ""
    part_no: str = ""
    part_name: str = ""
    supplier: str = ""
    engineer: str = ""
    test_date: str = ""
    theoretical_part_weight: float = 0.0  # g
    actual_part_weight: float = 0.0  # g
    
    # 模具信息
    mold_number: str = ""
    runner_type: str = ""
    cavity_count: str = ""  # 例如 "4" 或 "1+1"
    mold_size: str = ""  # LWH (mm)
    gate_type: str = ""
    
    # 材料信息
    material_brand: str = ""
    material_grade: str = ""
    material_number: str = ""
    material_color: str = ""
    material_density: float = 0.0  # g/cm³
    drying_temp: str = ""  # °C
    drying_time: str = ""  # H
    recommended_mold_temp: str = ""  # °C
    recommended_melt_temp: str = ""  # °C
    mfr: str = ""  # g/10min
    
    # 机台信息
    machine_brand: str = ""
    machine_model: str = ""
    machine_number: str = ""
    machine_type: str = ""  # 油压机/电动机
    machine_tonnage: float = 0.0  # 吨
    screw_diameter: float = 0.0  # mm
    intensification_ratio: float = 0.0
    retention_time: float = 0.0  # min
    shot_percentage: float = 0.0  # %
    cycle_time: float = 0.0  # s
    
    # 工艺参数
    barrel_temp_zone1: float = 0.0
    barrel_temp_zone2: float = 0.0
    barrel_temp_zone3: float = 0.0
    barrel_temp_zone4: float = 0.0  # Zone 4
    barrel_temp_zone5: float = 0.0  # Zone 5
    nozzle_temp: float = 0.0
    hot_runner_temp: float = 0.0   # Added
    mold_temp_fixed: float = 0.0
    mold_temp_moving: float = 0.0
    max_injection_pressure: float = 0.0
    max_holding_pressure: float = 0.0
    vp_transfer_position: float = 0.0


@dataclass
class ExcelTestData:
    """完整的Excel测试数据"""
    # 基本信息
    project_name: str = ""
    mold_name: str = ""
    material_name: str = ""
    machine_name: str = ""
    operator: str = ""
    test_date: str = ""
    
    # 机台快照
    machine_snapshot: Optional[MachineSnapshotData] = None
    
    # 7步数据
    viscosity: Optional[ViscosityData] = None
    cavity_balance: Optional[CavityBalanceData] = None
    pressure_drop: Optional[PressureDropData] = None
    process_window: Optional[ProcessWindowData] = None
    gate_freeze: Optional[GateFreezeData] = None
    cooling_time: Optional[CoolingTimeData] = None
    clamping_force: Optional[ClampingForceData] = None
    
    # 解析状态
    parse_errors: List[str] = field(default_factory=list)
    parse_warnings: List[str] = field(default_factory=list)


class ExcelDataParser:
    """Excel数据解析器"""
    
    def __init__(self):
        self.result = ExcelTestData()
    
    def parse_file(self, file_path: str) -> ExcelTestData:
        """从文件路径解析Excel"""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            return self._parse_workbook(wb)
        except Exception as e:
            self.result.parse_errors.append(f"无法打开文件: {str(e)}")
            return self.result
    
    def parse_bytes(self, file_content: bytes) -> ExcelTestData:
        """从字节流解析Excel（用于上传）"""
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            return self._parse_workbook(wb)
        except Exception as e:
            self.result.parse_errors.append(f"无法解析文件: {str(e)}")
            return self.result
    
    def _parse_workbook(self, wb: Workbook) -> ExcelTestData:
        """解析工作簿"""
        self.result = ExcelTestData()
        
        # 获取所有工作表名称
        sheet_names = wb.sheetnames
        
        # 智能匹配工作表
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            sheet_lower = sheet_name.lower()
            
            # 尝试识别工作表类型
            if any(k in sheet_lower for k in ['粘度', 'viscosity', 'step1', '步骤1']):
                self._parse_viscosity_sheet(ws)
            elif any(k in sheet_lower for k in ['型腔', 'cavity', 'balance', 'step2', '步骤2']):
                self._parse_cavity_balance_sheet(ws)
            elif any(k in sheet_lower for k in ['压力', 'pressure', 'drop', 'step3', '步骤3']):
                self._parse_pressure_drop_sheet(ws)
            elif any(k in sheet_lower for k in ['工艺窗口', 'process', 'window', 'step4', '步骤4']):
                self._parse_process_window_sheet(ws)
            elif any(k in sheet_lower for k in ['浇口', 'gate', 'freeze', 'seal', 'step5', '步骤5']):
                self._parse_gate_freeze_sheet(ws)
            elif any(k in sheet_lower for k in ['冷却', 'cooling', 'step6', '步骤6']):
                self._parse_cooling_time_sheet(ws)
            elif any(k in sheet_lower for k in ['锁模', 'clamping', 'force', 'step7', '步骤7']):
                self._parse_clamping_force_sheet(ws)
            elif any(k in sheet_lower for k in ['机台', 'machine', '参数', 'parameter', 'info', '信息']):
                self._parse_machine_info_sheet(ws)
            elif any(k in sheet_lower for k in ['数据', 'data', '测试', 'test']):
                # 通用数据表，尝试自动识别
                self._parse_generic_data_sheet(ws)
        
        # 如果没有找到任何有效数据，尝试解析第一个工作表
        if not self._has_any_data():
            if sheet_names:
                self._parse_generic_data_sheet(wb[sheet_names[0]])
        
        return self.result
    
    def _has_any_data(self) -> bool:
        """检查是否已解析到任何数据"""
        return any([
            self.result.viscosity and self.result.viscosity.speeds,
            self.result.cavity_balance and self.result.cavity_balance.cavity_weights,
            self.result.pressure_drop and self.result.pressure_drop.pressures,
            self.result.process_window and self.result.process_window.speeds,
            self.result.gate_freeze and self.result.gate_freeze.hold_times,
            self.result.cooling_time and self.result.cooling_time.cooling_times,
            self.result.clamping_force and self.result.clamping_force.forces,
        ])
    
    def _parse_viscosity_sheet(self, ws):
        """解析粘度曲线工作表"""
        data = ViscosityData()
        
        # 查找数据区域
        speeds = []
        viscosities = []
        
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_value = str(cell.value or "").lower()
                
                # 查找表头和数据
                if any(k in cell_value for k in ['射速', 'speed', '速度', 'mm/s']):
                    # 找到射速列，读取下面的数据
                    speeds = self._read_column_numbers(ws, row + 1, col)
                elif any(k in cell_value for k in ['粘度', 'viscosity', 'mpa']):
                    viscosities = self._read_column_numbers(ws, row + 1, col)
                elif any(k in cell_value for k in ['螺杆', 'screw', '直径']):
                    # 查找螺杆直径值
                    next_cell = ws.cell(row=row, column=col + 1).value
                    if next_cell and self._is_number(next_cell):
                        data.screw_diameter = float(next_cell)
                elif any(k in cell_value for k in ['材料', 'material']):
                    next_cell = ws.cell(row=row, column=col + 1).value
                    if next_cell:
                        data.material = str(next_cell)
        
        # 如果没找到标题，尝试智能识别两列数字
        if not speeds or not viscosities:
            speeds, viscosities = self._find_two_number_columns(ws)
        
        if speeds and viscosities:
            # 确保长度一致
            min_len = min(len(speeds), len(viscosities))
            data.speeds = speeds[:min_len]
            data.viscosities = viscosities[:min_len]
            self.result.viscosity = data
        else:
            self.result.parse_warnings.append("粘度工作表: 未找到有效数据")
    
    def _parse_cavity_balance_sheet(self, ws):
        """解析型腔平衡工作表"""
        data = CavityBalanceData()
        
        headers = {}
        data_row_start = 2
        
        # 查找表头
        for row in range(1, 5):
            for col in range(1, 10):
                val = str(ws.cell(row=row, column=col).value or "").lower()
                if any(k in val for k in ['腔', 'cavity']):
                    headers['cavity'] = col
                    data_row_start = row + 1
                elif any(k in val for k in ['重量', 'weight']):
                    headers['weight'] = col
                elif any(k in val for k in ['判定', 'check', 'visual']):
                    headers['visual'] = col
                elif any(k in val for k in ['类型', 'type']):
                    headers['type'] = col

        if 'cavity' in headers and 'weight' in headers:
            for row in range(data_row_start, ws.max_row + 1):
                cav_val = ws.cell(row=row, column=headers['cavity']).value
                weight_val = ws.cell(row=row, column=headers['weight']).value
                
                if cav_val is not None and self._is_number(cav_val):
                    cav_idx = int(float(cav_val))
                    if weight_val is not None and self._is_number(weight_val):
                        data.cavity_weights[cav_idx] = float(weight_val)
                    
                    if 'visual' in headers:
                        vis_val = ws.cell(row=row, column=headers['visual']).value
                        if vis_val:
                            data.visual_checks[cav_idx] = str(vis_val)
        
        if data.cavity_weights:
            self.result.cavity_balance = data
    
    def _parse_pressure_drop_sheet(self, ws):
        """解析压力降工作表"""
        data = PressureDropData()
        
        positions = []
        pressures = []
        
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_value = str(cell.value or "").lower()
                
                if any(k in cell_value for k in ['位置', 'position', 'location']):
                    positions = self._read_column_strings(ws, row + 1, col)
                elif any(k in cell_value for k in ['压力', 'pressure', 'mpa']):
                    pressures = self._read_column_numbers(ws, row + 1, col)
        
        if positions and pressures:
            min_len = min(len(positions), len(pressures))
            data.positions = positions[:min_len]
            data.pressures = pressures[:min_len]
            self.result.pressure_drop = data
    
    def _parse_process_window_sheet(self, ws):
        """解析工艺窗口工作表 (Step 4)"""
        data = ProcessWindowData()
        
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_value = str(cell.value or "").lower()
                
                if any(k in cell_value for k in ['射速', 'speed', '速度']):
                    data.speeds = self._read_column_numbers(ws, row + 1, col)
                elif any(k in cell_value for k in ['压力', 'pressure']):
                    data.pressures = self._read_column_numbers(ws, row + 1, col)
                elif any(k in cell_value for k in ['产品重量', 'product height', 'weight']):
                    data.product_weights = self._read_column_numbers(ws, row + 1, col)
                elif any(k in cell_value for k in ['合格', 'ok', 'pass', 'quality']):
                    data.quality_ok = self._read_column_bools(ws, row + 1, col)
        
        if data.speeds and data.pressures:
            self.result.process_window = data
    
    def _parse_gate_freeze_sheet(self, ws):
        """解析浇口冻结工作表"""
        data = GateFreezeData()
        
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_value = str(cell.value or "").lower()
                
                if any(k in cell_value for k in ['保压时间', 'hold', 'time', '时间']):
                    data.hold_times = self._read_column_numbers(ws, row + 1, col)
                elif any(k in cell_value for k in ['重量', 'weight', '克', 'gram']):
                    data.weights = self._read_column_numbers(ws, row + 1, col)
        
        if data.hold_times and data.weights:
            self.result.gate_freeze = data
    
    def _parse_cooling_time_sheet(self, ws):
        """解析冷却时间工作表"""
        data = CoolingTimeData()
        
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_value = str(cell.value or "").lower()
                
                if any(k in cell_value for k in ['冷却时间', 'cooling', '秒']):
                    data.cooling_times = self._read_column_numbers(ws, row + 1, col)
                elif any(k in cell_value for k in ['温度', 'temp', '°c']):
                    data.part_temps = self._read_column_numbers(ws, row + 1, col)
                elif any(k in cell_value for k in ['变形', 'deform', 'mm']):
                    data.deformations = self._read_column_numbers(ws, row + 1, col)
        
        if data.cooling_times:
            self.result.cooling_time = data
    
    def _parse_clamping_force_sheet(self, ws):
        """解析锁模力工作表"""
        data = ClampingForceData()
        
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_value = str(cell.value or "").lower()
                
                if any(k in cell_value for k in ['锁模力', 'clamp', 'force', '吨']):
                    data.forces = self._read_column_numbers(ws, row + 1, col)
                elif any(k in cell_value for k in ['飞边', 'flash', '溢料']):
                    data.flash_detected = self._read_column_bools(ws, row + 1, col)
        
        if data.forces:
            self.result.clamping_force = data
    
    def _parse_machine_info_sheet(self, ws):
        """解析机台信息工作表"""
        data = MachineSnapshotData()
        
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, min(30, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_value = str(cell.value or "").lower()
                next_cell = ws.cell(row=row, column=col + 1).value
                
                # 有些是数字，有些是字符串
                if not next_cell:
                    continue
                
                is_num = self._is_number(next_cell)
                value = float(next_cell) if is_num else 0.0
                
                if any(k in cell_value for k in ['料筒1', 'barrel1', '一段']):
                    data.barrel_temp_zone1 = value
                elif any(k in cell_value for k in ['料筒2', 'barrel2', '二段']):
                    data.barrel_temp_zone2 = value
                elif any(k in cell_value for k in ['料筒3', 'barrel3', '三段']):
                    data.barrel_temp_zone3 = value
                elif any(k in cell_value for k in ['料筒4', 'barrel4', '四段']):
                    data.barrel_temp_zone4 = value
                elif any(k in cell_value for k in ['料筒5', 'barrel5', '五段']):
                    data.barrel_temp_zone5 = value
                elif any(k in cell_value for k in ['射嘴', 'nozzle']):
                    data.nozzle_temp = value
                elif any(k in cell_value for k in ['热流道', 'hot runner']):
                    data.hot_runner_temp = value
                elif any(k in cell_value for k in ['成型周期', 'cycle time']):
                    data.cycle_time = value
                elif any(k in cell_value for k in ['定模', 'fixed']):
                    data.mold_temp_fixed = value
                elif any(k in cell_value for k in ['动模', 'moving']):
                    data.mold_temp_moving = value
                elif any(k in cell_value for k in ['项目', 'project']):
                    self.result.project_name = str(next_cell)
                elif any(k in cell_value for k in ['模具', 'mold']):
                    self.result.mold_name = str(next_cell)
                elif any(k in cell_value for k in ['材料', 'material']):
                    self.result.material_name = str(next_cell)
                elif any(k in cell_value for k in ['机台', 'machine']):
                    self.result.machine_name = str(next_cell)
        
        if any([data.barrel_temp_zone1, data.mold_temp_fixed]):
            self.result.machine_snapshot = data
    
    def _parse_generic_data_sheet(self, ws):
        """通用数据表解析 - 尝试自动识别数据"""
        # 扫描所有单元格，查找关键字
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_value = str(cell.value or "").lower()
                
                # 根据关键字触发相应解析
                if any(k in cell_value for k in ['粘度', 'viscosity']):
                    self._parse_viscosity_sheet(ws)
                    return
                elif any(k in cell_value for k in ['型腔', 'cavity']):
                    self._parse_cavity_balance_sheet(ws)
                    return
    
    def _read_column_numbers(self, ws, start_row: int, col: int, max_rows: int = 50) -> List[float]:
        """读取一列数字"""
        numbers = []
        for row in range(start_row, min(start_row + max_rows, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None and self._is_number(cell_value):
                numbers.append(float(cell_value))
            elif cell_value is None or str(cell_value).strip() == "":
                # 空行，停止读取
                if numbers:  # 已有数据时才停止
                    break
        return numbers
    
    def _read_column_strings(self, ws, start_row: int, col: int, max_rows: int = 50) -> List[str]:
        """读取一列字符串"""
        strings = []
        for row in range(start_row, min(start_row + max_rows, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None and str(cell_value).strip():
                strings.append(str(cell_value).strip())
            elif not strings:
                continue
            else:
                break
        return strings
    
    def _read_column_bools(self, ws, start_row: int, col: int, max_rows: int = 50) -> List[bool]:
        """读取一列布尔值"""
        bools = []
        for row in range(start_row, min(start_row + max_rows, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                val_str = str(cell_value).lower()
                is_true = val_str in ['true', 'yes', '是', '合格', 'ok', '1', 'pass', '√', '✓']
                bools.append(is_true)
            else:
                if bools:
                    break
        return bools
    
    def _find_two_number_columns(self, ws) -> Tuple[List[float], List[float]]:
        """智能查找两列数字数据（用于粘度曲线）"""
        # 扫描找到第一个连续数字列
        number_columns = []
        
        for col in range(1, min(20, ws.max_column + 1)):
            numbers = []
            for row in range(1, min(100, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None and self._is_number(cell_value):
                    numbers.append((row, float(cell_value)))
            
            # 如果这一列有连续的数字（至少3个）
            if len(numbers) >= 3:
                number_columns.append((col, numbers))
        
        # 返回前两列数字
        if len(number_columns) >= 2:
            col1_data = [v for _, v in number_columns[0][1]]
            col2_data = [v for _, v in number_columns[1][1]]
            return col1_data, col2_data
        
        return [], []
    
    def _is_number(self, value) -> bool:
        """检查值是否为数字"""
        if isinstance(value, (int, float)):
            return True
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False


def create_template_excel(output_path: str) -> str:
    """创建数据输入模板Excel文件"""
    wb = openpyxl.Workbook()
    
    # Step 1: 粘度曲线 - 原始测量数据
    ws1 = wb.active
    ws1.title = "Step1_粘度曲线"
    ws1['A1'] = "速度 (%)"
    ws1['B1'] = "实际速度 (mm/s)"
    ws1['C1'] = "切换位置 (mm)"
    ws1['D1'] = "填充时间 (s)"
    ws1['E1'] = "峰值压力 (Bar)"
    ws1['G1'] = "螺杆直径:"
    ws1['H1'] = 53
    ws1['G2'] = "材料:"
    ws1['H2'] = "PA6 GF30"
    ws1['G3'] = "说明:"
    ws1['H3'] = "填写原始测量值，App将计算粘度"
    
    # 示例数据 - 来自seed_data.py
    from seed_data import ScientificMoldingSeedData
    seed_gen = ScientificMoldingSeedData(seed=42)
    visc_data = seed_gen.generate_viscosity_raw_data()
    
    for i, point in enumerate(visc_data, start=2):
        ws1[f'A{i}'] = point['speed_percent']
        ws1[f'B{i}'] = point['speed_mm_s']
        ws1[f'C{i}'] = point['switch_position']
        ws1[f'D{i}'] = point['fill_time']
        ws1[f'E{i}'] = point['peak_pressure']
    
    # Step 2: 型腔平衡 - 区分短射和满射
    ws2 = wb.create_sheet("Step2_型腔平衡")
    ws2['A1'] = "测试类型"
    ws2['B1'] = "腔号"
    ws2['C1'] = "重量 (g)"
    ws2['D1'] = "目视判定 (OK/NG)"  # Added
    ws2['F1'] = "说明:"
    ws2['G1'] = "Short_Shot=短射50%, VP_Switch=满射99%"
    
    # 短射数据
    row_idx = 2
    bal_data = seed_gen.generate_cavity_balance_data(num_cavities=8)
    for point in bal_data['short_shot']:
        ws2[f'A{row_idx}'] = "Short_Shot"
        ws2[f'B{row_idx}'] = point['cavity_index']
        ws2[f'C{row_idx}'] = point['weight']
        ws2[f'D{row_idx}'] = point['visual_check']  # Added
        row_idx += 1
    
    # 满射数据
    for point in bal_data['vp_switch']:
        ws2[f'A{row_idx}'] = "VP_Switch"
        ws2[f'B{row_idx}'] = point['cavity_index']
        ws2[f'C{row_idx}'] = point['weight']
        ws2[f'D{row_idx}'] = point['visual_check']  # Added
        row_idx += 1
    
    # Step 3: 压力降测试 - 标准测量位置（枚举值）
    ws3 = wb.create_sheet("Step3_压力降")
    ws3['A1'] = "位置"
    ws3['B1'] = "压力 (Bar)"
    ws3['D1'] = "说明:"
    ws3['E1'] = "位置必须是: Nozzle, Runner, Gate, Part_50%, Part_99%"
    
    pressure_data = seed_gen.generate_pressure_drop_data()
    for i, point in enumerate(pressure_data, start=2):
        ws3[f'A{i}'] = point['position']
        ws3[f'B{i}'] = point['pressure']
    
    # Step 4: 工艺窗口
    ws4 = wb.create_sheet("Step4_工艺窗口")
    ws4['A1'] = "射速 (mm/s)"
    ws4['B1'] = "保压压力 (Bar)"
    ws4['C1'] = "产品重量 (g)"  # Added
    ws4['D1'] = "保压时间 (s)"
    ws4['E1'] = "产品质量"
    ws4['G1'] = "说明:"
    ws4['H1'] = "质量: Pass/Fail"
    
    window_data = seed_gen.generate_process_window_data()
    for i, point in enumerate(window_data, start=2):
        ws4[f'A{i}'] = point['speed_mm_s']
        ws4[f'B{i}'] = point['hold_pressure_bar']
        ws4[f'C{i}'] = point['product_weight']  # Added
        ws4[f'D{i}'] = point['hold_time']
        ws4[f'E{i}'] = point['quality']
    
    # Step 5: 浇口冻结
    ws5 = wb.create_sheet("Step5_浇口冻结")
    ws5['A1'] = "保压时间 (s)"
    ws5['B1'] = "重量 (g)"
    ws5['D1'] = "说明:"
    ws5['E1'] = "逐步增加保压时间，记录产品重量变化"
    
    freeze_data = seed_gen.generate_gate_freeze_data()
    for i, point in enumerate(freeze_data, start=2):
        ws5[f'A{i}'] = point['hold_time']
        ws5[f'B{i}'] = point['weight']
    
    # Step 6: 冷却时间
    ws6 = wb.create_sheet("Step6_冷却时间")
    ws6['A1'] = "冷却时间 (s)"
    ws6['B1'] = "产品温度 (°C)"
    ws6['C1'] = "变形量 (mm)"
    ws6['E1'] = "说明:"
    ws6['F1'] = "测试不同冷却时间对产品质量的影响"
    
    cooling_data = seed_gen.generate_cooling_time_data()
    for i, point in enumerate(cooling_data, start=2):
        ws6[f'A{i}'] = point['cooling_time']
        ws6[f'B{i}'] = point['part_temp']
        ws6[f'C{i}'] = point['deformation']
    
    # Step 7: 锁模力优化
    ws7 = wb.create_sheet("Step7_锁模力")
    ws7['A1'] = "锁模力 (吨)"
    ws7['B1'] = "产品重量 (g)"
    ws7['C1'] = "飞边情况"
    ws7['E1'] = "说明:"
    ws7['F1'] = "飞边情况: Yes/No，重量变化辅助判断"
    
    clamp_data = seed_gen.generate_clamping_force_data()
    for i, point in enumerate(clamp_data, start=2):
        ws7[f'A{i}'] = point['clamping_force']
        ws7[f'B{i}'] = point['part_weight']
        ws7[f'C{i}'] = point['flash_detected']
        ws7[f'B{i}'] = point['flash_detected']
    
    # 项目综合信息 - MIL标准格式（扩充版）
    ws_project = wb.create_sheet("项目综合信息")
    ws_project['A1'] = "参数分类"
    ws_project['B1'] = "参数名称"
    ws_project['C1'] = "参数值"
    
    # 获取完整测试套件数据
    suite = seed_gen.generate_complete_test_suite()
    
    project_params = [
        ("=== 产品信息 ===", "", ""),
        ("产品信息", "Model No", suite['project_info']['model_no']),
        ("产品信息", "Part No", suite['project_info']['part_no']),
        ("产品信息", "Part Name", suite['project_info']['part_name']),
        ("产品信息", "供应商 Supplier", suite['project_info']['supplier']),
        ("产品信息", "负责人 Engineer", suite['project_info']['engineer']),
        ("产品信息", "测试日期", suite['project_info']['test_date']),
        ("产品信息", "理论重量 (g)", suite['part_info']['theoretical_weight']),
        ("产品信息", "实际重量 (g)", suite['part_info']['actual_weight']),
        
        ("=== 模具信息 ===", "", ""),
        ("模具信息", "模号 Mold Number", suite['mold_info']['mold_number']),
        ("模具信息", "流道形式 Runner Type", suite['mold_info']['runner_type']),
        ("模具信息", "模穴数 Cavity Qty", suite['mold_info']['cavity_count']),
        ("模具信息", "模具尺寸 Mold Size", suite['mold_info']['mold_size']),
        ("模具信息", "浇口类型 Gate Type", suite['mold_info']['gate_type']),
        
        ("=== 材料信息 ===", "", ""),
        ("材料信息", "品牌 Brand", suite['material_info']['brand']),
        ("材料信息", "型号 Grade", suite['material_info']['grade']),
        ("材料信息", "材料编号", suite['material_info']['material_number']),
        ("材料信息", "颜色 Color", suite['material_info']['color']),
        ("材料信息", "密度 Density (g/cm³)", suite['material_info']['density']),
        ("材料信息", "烘烤温度 (°C)", suite['material_info']['drying_temp']),
        ("材料信息", "烘烤时间 (H)", suite['material_info']['drying_time']),
        ("材料信息", "推荐模温 (°C)", suite['material_info']['recommended_mold_temp']),
        ("材料信息", "推荐料温 (°C)", suite['material_info']['recommended_melt_temp']),
        ("材料信息", "MFR (g/10min)", suite['material_info']['mfr']),
        
        ("=== 机台信息 ===", "", ""),
        ("机台信息", "品牌 Brand", suite['machine_info']['brand']),
        ("机台信息", "型号 Model", suite['machine_info']['model']),
        ("机台信息", "机台号 Machine #", suite['machine_info']['machine_number']),
        ("机台信息", "类型 Type", suite['machine_info']['machine_type']),
        ("机台信息", "吨位 Tonnage", suite['machine_info']['tonnage']),
        ("机台信息", "螺杆直径 (mm)", suite['machine_info']['screw_diameter']),
        ("机台信息", "增压比 Intensification Ratio", suite['machine_info']['intensification_ratio']),
        ("机台信息", "滞留时间 (min)", suite['machine_info']['retention_time']),
        ("机台信息", "占总胶量百分比 (%)", suite['machine_info']['shot_percentage']),
        ("机台信息", "周期时间 Cycle Time (s)", suite['machine_info']['cycle_time']),
        
        ("=== 工艺参数 ===", "", ""),
        ("工艺参数", "料筒温度-1段 Zone 1 (°C)", suite['machine_info']['barrel_temps'][0]),
        ("工艺参数", "料筒温度-2段 Zone 2 (°C)", suite['machine_info']['barrel_temps'][1]),
        ("工艺参数", "料筒温度-3段 Zone 3 (°C)", suite['machine_info']['barrel_temps'][2]),
        ("工艺参数", "料筒温度-4段 Zone 4 (°C)", suite['machine_info']['barrel_temps'][3]),
        ("工艺参数", "料筒温度-5段 Zone 5 (°C)", suite['machine_info']['barrel_temps'][4]),
        ("工艺参数", "射嘴温度 Nozzle (°C)", suite['material_info']['recommended_melt_temp']),
        ("工艺参数", "热流道温度 Hot Runner (°C)", suite['machine_info']['hot_runner_temp']),
        ("工艺参数", "模温-定模 Fixed Mold (°C)", suite['material_info']['recommended_mold_temp']),
        ("工艺参数", "模温-动模 Moving Mold (°C)", suite['material_info']['recommended_mold_temp']),
        ("工艺参数", "V/P切换位置 (mm)", suite['machine_info']['vp_switch_position']),
        ("工艺参数", "最大压力 (Bar)", suite['machine_info']['max_pressure_bar']),
        ("工艺参数", "最大射速 (mm/s)", suite['machine_info']['max_speed_mm_s']),
    ]
    
    for i, (category, name, value) in enumerate(project_params, start=2):
        ws_project[f'A{i}'] = category
        ws_project[f'B{i}'] = name
        ws_project[f'C{i}'] = value
    
    # 添加说明工作表
    ws_info = wb.create_sheet("使用说明", 0)
    ws_info['A1'] = "SmartMold 数据上传模板 - 使用说明"
    ws_info['A1'].font = openpyxl.styles.Font(size=16, bold=True, color="0000FF")
    
    instructions = [
        "",
        "📋 本模板包含科学注塑7步法的所有测试数据（原始测量值）:",
        "",
        "• Step1_粘度曲线: 速度%、实际速度、填充时间、峰值压力",
        "  → App将计算: 剪切率、有效粘度",
        "",
        "• Step2_型腔平衡: 测试类型（Short_Shot/VP_Switch）、腔号、重量",
        "  → App将计算: 平衡度、不合格腔",
        "",
        "• Step3_压力降: 标准位置（Nozzle/Runner/Gate/Part_50%/Part_99%）、压力",
        "  → App将计算: 压降梯度、阻力分析",
        "",
        "• Step4_工艺窗口: 射速、保压压力、产品质量（Pass/Fail）",
        "  → App将计算: 工艺窗口边界",
        "",
        "• Step5_浇口冻结: 保压时间、重量",
        "  → App将计算: 浇口冻结时间点",
        "",
        "• Step6_冷却时间: 冷却时间、产品温度、变形量",
        "  → App将计算: 最佳冷却时间",
        "",
        "• Step7_锁模力: 锁模力、飞边情况（Yes/No）",
        "  → App将计算: 最小安全锁模力",
        "",
        "• 机台参数: 必须包含螺杆直径、增压比、最大压力（用于计算）",
        "",
        "⚙️ 核心原则: App是计算器，不是记录本",
        "",
        "✓ 您提供: 原始测量数据（Raw Data）",
        "✓ App计算: 派生指标（Insight）",
        "",
        "✏️ 使用方法:",
        "",
        "1. 填写您的实际测量数据（不要自己算粘度！）",
        "2. 可以只填写部分步骤",
        "3. 保持表头格式不变",
        "4. 在SmartMold系统上传",
        "5. 系统自动计算并生成报告",
        "",
        "💡 重要提示:",
        "",
        "• Step3位置必须用枚举值: Nozzle, Runner, Gate, Part_50%, Part_99%",
        "• Step2测试类型必须是: Short_Shot 或 VP_Switch",
        "• 机台参数中的螺杆直径、增压比是必填项",
        "• 所有示例数据都符合物理规律，可作为参考",
    ]
    
    for i, text in enumerate(instructions, start=2):
        ws_info[f'A{i}'] = text
        if text.startswith(("•", "✏️", "💡", "📋")):
            ws_info[f'A{i}'].font = openpyxl.styles.Font(bold=True)
    
    # 调整列宽
    ws_info.column_dimensions['A'].width = 60
    
    wb.save(output_path)
    return output_path


# 测试代码
if __name__ == "__main__":
    # 创建模板
    template_path = "test_data_template.xlsx"
    create_template_excel(template_path)
    print(f"✓ 模板已创建: {template_path}")
    
    # 测试解析
    parser = ExcelDataParser()
    result = parser.parse_file(template_path)
    
    print(f"\n解析结果:")
    if result.viscosity:
        print(f"  粘度数据: {len(result.viscosity.speeds)} 个点")
        print(f"    射速: {result.viscosity.speeds}")
        print(f"    粘度: {result.viscosity.viscosities}")
    if result.cavity_balance:
        print(f"  型腔平衡: {len(result.cavity_balance.cavity_weights)} 腔")
    if result.gate_freeze:
        print(f"  浇口冻结: {len(result.gate_freeze.hold_times)} 个点")
    if result.cooling_time:
        print(f"  冷却时间: {len(result.cooling_time.cooling_times)} 个点")
    
    if result.parse_errors:
        print(f"\n错误: {result.parse_errors}")
    if result.parse_warnings:
        print(f"\n警告: {result.parse_warnings}")
